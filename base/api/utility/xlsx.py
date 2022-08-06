from opcode import opname
from webbrowser import get
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import string
import random

border = Border(left=Side(border_style='thin',
                        color='FF000000'),
              right=Side(border_style='thin',
                         color='FF000000'),
              top=Side(border_style='thin',
                       color='FF000000'),
              bottom=Side(border_style='thin',
                          color='FF000000'),
              diagonal=Side(border_style='thin',
                            color='FF000000'),
              diagonal_direction=0,
              outline=Side(border_style='thin',
                           color='FF000000'),
              vertical=Side(border_style='thin',
                            color='FF000000'),
              horizontal=Side(border_style='thin',
                             color='FF000000')
             )

grayFill = PatternFill(fill_type='solid',
start_color='bdbdbd',
end_color='bdbdbd')

lightGrayFill = PatternFill(fill_type='solid',
start_color='f0f0f0',
end_color='f0f0f0')

titleFont = Font(name='Calibri',
            size=14,
            bold=True,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')


def xlsx_paraescolares(students, paraescolar, turno):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['B2'] = 'Paraescolar de ' + paraescolar
    ws['B2'].font = titleFont
    ws['B3'] = 'Turno ' + turno
    ws['B3'].font = titleFont

    ws['A5'].fill = grayFill
    ws['B5'].fill = grayFill
    ws['C5'].fill = grayFill
    ws['D5'].fill = grayFill

    ws['A5'].border = border
    ws['B5'].border = border
    ws['C5'].border = border
    ws['D5'].border = border


    ws['B5'] = 'Matrícula'
    ws['C5'] = 'Nombre'
    ws['D5'] = 'Grupo'

    students = sorted(students, key=lambda x : x['nombre_completo'])

    cont = 0
    for row in ws.iter_rows(min_row=6, min_col=1, max_col=4, max_row=5 + len(students)):
        row[0].fill = lightGrayFill
        row[1].fill = lightGrayFill
        row[2].fill = lightGrayFill
        row[3].fill = lightGrayFill

        row[0].border = border
        row[1].border = border
        row[2].border = border
        row[3].border = border

        row[0].value = cont + 1
        row[1].value = students[cont]['matricula']
        row[2].value = students[cont]['nombre_completo']
        row[3].value = students[cont]['grupo']
        cont += 1


    filename = paraescolar + '-' + turno + '.xlsx'
    filepath = 'base/api/utility/' + filename
    wb.save(filename=filepath)

    return filename



def xlsx_grupos(students, grupo):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['B2'] = 'Grupo ' + str(grupo)
    ws['B2'].font = titleFont

    ws['A4'].fill = grayFill
    ws['B4'].fill = grayFill
    ws['C4'].fill = grayFill
    ws['D4'].fill = grayFill

    ws['A4'].border = border
    ws['B4'].border = border
    ws['C4'].border = border
    ws['D4'].border = border


    ws['B4'] = 'Matrícula'
    ws['C4'] = 'Nombre'
    ws['D4'] = 'Paraescolar'

    students = sorted(students, key=lambda x : x['nombre_completo'])

    cont = 0
    for row in ws.iter_rows(min_row=5, min_col=1, max_col=4, max_row=4 + len(students)):
        row[0].fill = lightGrayFill
        row[1].fill = lightGrayFill
        row[2].fill = lightGrayFill
        row[3].fill = lightGrayFill

        row[0].border = border
        row[1].border = border
        row[2].border = border
        row[3].border = border

        row[0].value = cont + 1
        row[1].value = students[cont]['matricula']
        row[2].value = students[cont]['nombre_completo']
        row[3].value = students[cont]['paraescolar']
        cont += 1


    filename = grupo + '.xlsx'
    filepath = 'base/api/utility/' + filename
    wb.save(filename=filepath)
    #print('SAVED ' + filename)

    return filename


border = Border(left=Side(border_style='thin',
                        color='FF000000'),
              right=Side(border_style='thin',
                         color='FF000000'),
              top=Side(border_style='thin',
                       color='FF000000'),
              bottom=Side(border_style='thin',
                          color='FF000000'),
              diagonal=Side(border_style='thin',
                            color='FF000000'),
              diagonal_direction=0,
              outline=Side(border_style='thin',
                           color='FF000000'),
              vertical=Side(border_style='thin',
                            color='FF000000'),
              horizontal=Side(border_style='thin',
                             color='FF000000')
             )


if __name__ == '__main__':
    #wb = openpyxl.Workbook()
    # sh = wb.active

    # x = sh['A1']
    # x.border = border
    # wb.save('test.xlsx')
    students = [{'id': 8, 'nombre_completo': 'PANCHO PONCHO TRIJO TRAJO', 'grupo': '452', 'paraescolar': None, 'matricula': '19053482', 'turno': 'VESPERTINO', 'tiene_paraescolar': False}, {'id': 9, 'nombre_completo': 'PEREZ SANCHEZ JUAN PABLO', 'grupo': '356', 'paraescolar': None, 'matricula': '19392308', 'turno': 'VESPERTINO', 'tiene_paraescolar': False}, {'id': 7, 'nombre_completo': 'TAPIA LOYA SAUL ALEJANDRO', 'grupo': '609', 'paraescolar': None, 'matricula': '19080045', 'turno': 'MATUTINO', 'tiene_paraescolar': False}]
    xlsx_paraescolares(students, 'robotica', 'matutino')
    #xlsx_grupos(students, 609)